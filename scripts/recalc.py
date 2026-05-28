"""
Open an xlsx and check for formula errors (#REF!, #NAME?, #VALUE!, etc.).

Run after build_workbook.py and before committing the batch. The build_workbook.py
output doesn't typically use formulas, so this is mostly a defensive sanity
check — but the SOP requires it.

Usage:
    python recalc.py batches/2026-05-27_ref_fill_rows_1148-1167/lng_carrier_backend_ref_fill.xlsx

Exit code:
    0 = no errors found
    1 = errors found (printed to stderr)
"""
import sys
from pathlib import Path

from openpyxl import load_workbook


ERROR_TOKENS = ("#REF!", "#NAME?", "#VALUE!", "#DIV/0!", "#N/A", "#NULL!", "#NUM!")


def check_workbook(xlsx_path: str) -> tuple[int, int]:
    """Returns (sheets_checked, errors_found)."""
    wb = load_workbook(xlsx_path, data_only=False)
    sheets_checked = 0
    errors_found = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheets_checked += 1
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                v = str(cell.value)
                for tok in ERROR_TOKENS:
                    if tok in v:
                        print(f"  [ERROR] {sheet_name}!{cell.coordinate}: {v!r}",
                              file=sys.stderr)
                        errors_found += 1
                        break
    return sheets_checked, errors_found


def main():
    if len(sys.argv) != 2:
        print("Usage: python recalc.py <xlsx_path>")
        sys.exit(2)
    path = sys.argv[1]
    if not Path(path).exists():
        print(f"File not found: {path}", file=sys.stderr)
        sys.exit(2)
    sheets, errors = check_workbook(path)
    print(f"  Checked {sheets} sheet(s) in {path}")
    if errors:
        print(f"  Found {errors} formula error(s) — see stderr above")
        sys.exit(1)
    print(f"  Zero formula errors")
    sys.exit(0)


if __name__ == "__main__":
    main()
