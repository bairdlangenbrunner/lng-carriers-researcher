"""
Build the rolling output xlsx for either [ref]-fill or discovery workflows.

Per [ref]-Fill SOP §2 and Discovery SOP §5, output structure:

  ref_fill mode -> <out>/lng_carrier_backend_ref_fill.xlsx
    Sheets: README, backend_ref_fill, QA_review

  discovery mode -> <out>/lng_carrier_candidate_vessels.xlsx
    Sheets: README, candidate_vessels, QA_review, backend_status_flags

The --out argument is the DIRECTORY the xlsx is written into; the filename is
fixed by mode. Normally this is the batch directory under batches/, e.g.
batches/2026-05-27_ref_fill_rows_1148-1167/.

Color conventions ([ref]-Fill SOP §2.2 / Discovery SOP §5.2):
  Green   = high confidence (multi-source or primary/regulatory + value verbatim)
  Yellow  = medium confidence (entity-level or contested)
  Red     = low confidence / review needed
  Peach   = override of pre-existing backend value (ref_fill only)
  Gray    = pre-existing backend value, untouched (ref_fill only)
  (No fill) = blank cell — see QA_review for why

This script provides scaffolding only — the citation/candidate data must be
supplied via a JSON input. The JSON schema is documented in the docstrings
below.

Usage:
    python build_workbook.py --mode ref_fill --rows 1170-1190 \\
        --citations citations.json \\
        --out ../batches/2026-05-27_ref_fill_rows_1170-1190/
    python build_workbook.py --mode discovery \\
        --candidates candidates.json \\
        --out ../batches/2026-05-27_discovery/

citations.json schema (ref_fill mode):
    {
      "batch_label": "Batch 3 — rows 1170-1190",
      "cells": [
        {
          "row_id": "1170",
          "field": "hull_ref",       # canonical column key from colmap
          "url": "https://...",
          "confidence": "G",          # G/Y/R/peach
          "note": "CSB Samsung yard page contains 'Samsung 2783'"
        },
        ...
      ],
      "qa_log": [
        {"row_id":"1170", "field":"hull_ref", "action":"filled", "confidence":"G",
         "url":"https://...", "note":"..."},
        ...
      ],
      "data_conflicts": [...],          # optional
      "candidate_data_fills": [...],    # optional, paired data + ref proposals
      "defects_corrected": [...],       # optional
      "verification_log": [...]         # optional
    }

candidates.json schema (discovery mode):
    {
      "gap_window": "2026-01-01 onward",
      "yards_swept": [...],
      "candidates": [
        {
          "cluster_id": "C1",
          "cluster_label": "MISC FSRU at Samsung HI",
          "confidence": "G",
          "discovery_notes": "...",
          "row_data": {            # cells keyed by EXACT backend header strings
            "Shipowner": "MISC Berhad",
            "Shipowner [ref]": "https://...",
            "Shipbuilder": "Samsung Heavy Industries",
            ...
          }
          # NOTE: do NOT put the 7 yard-location columns (Shipbuilder yard
          # country/area + its [ref] + the 5 Yard location cells) in row_data —
          # they are autofilled from an existing backend row for the same
          # shipbuilder (Discovery SOP §6.7), or left blank if the shipbuilder
          # is new. Multiple URLs in one [ref] cell join with ", " (SOP §4.15).
        },
        ...
      ],
      "backend_status_flags": [...],
      "verification_log": [...],
      "methodology_audit": {...}
    }
"""
import argparse
import csv
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

from paths import backend_csv_path, work_dir
from normalize import normalize_builder, normalize_owner
from lookups import (CONTROLLED_VOCAB, AMBIGUOUS, load_builder_facts,
                     YARD_FACT_COLS as YARD_LOCATION_COLS)


# Color conventions
FILL_GREEN = PatternFill("solid", fgColor="C6EFCE")    # high confidence
FILL_YELLOW = PatternFill("solid", fgColor="FFEB9C")   # medium
FILL_RED = PatternFill("solid", fgColor="FFC7CE")      # low
FILL_PEACH = PatternFill("solid", fgColor="FFD9B3")    # override
FILL_GRAY = PatternFill("solid", fgColor="EEEEEE")     # pre-existing
FILL_HEADER = PatternFill("solid", fgColor="1F4E78")   # header background

CONFIDENCE_FILLS = {
    "G": FILL_GREEN, "green": FILL_GREEN,
    "Y": FILL_YELLOW, "yellow": FILL_YELLOW,
    "R": FILL_RED, "red": FILL_RED,
    "peach": FILL_PEACH, "P": FILL_PEACH,
}

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
WRAP_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def _apply_header_style(ws, row_idx: int = 1):
    """Apply header styling to a row."""
    for cell in ws[row_idx]:
        cell.font = HEADER_FONT
        cell.fill = FILL_HEADER
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _resolve_out_path(out_arg: str | None, default_filename: str) -> Path:
    """
    Resolve --out to a concrete file path.

    Semantics:
      - If --out ends in .xlsx, treat it as a file path (backward compat).
      - If --out is a directory (existing or not), append default_filename.
      - If --out is None, default to <repo_root>/batches/_latest/<default_filename>
        and warn — the canonical pattern is to pass an explicit batch directory.
    """
    if out_arg is None:
        from paths import repo_root
        default = repo_root() / "batches" / "_latest" / default_filename
        print(f"  [warn] --out not specified; defaulting to {default}. "
              f"Pass --out <batches/...> for proper batch tracking.",
              file=sys.stderr)
        return default
    p = Path(out_arg)
    if p.suffix.lower() == ".xlsx":
        return p
    return p / default_filename


def _set_widths(ws, widths: dict):
    """widths: {column_letter: width}"""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def build_readme(wb: Workbook, title: str, body_lines: list[str]):
    ws = wb.active
    ws.title = "README"
    ws["A1"] = title
    ws["A1"].font = Font(name="Calibri", size=14, bold=True)
    for i, line in enumerate(body_lines, start=3):
        ws.cell(row=i, column=1, value=line).alignment = WRAP_ALIGN
    ws.column_dimensions["A"].width = 120


def build_ref_fill(args):
    """Build the [ref]-fill mode workbook."""
    citations_path = Path(args.citations)
    citations = json.loads(citations_path.read_text())

    # Load backend
    with open(args.backend, encoding="utf-8") as f:
        backend_rows = list(csv.reader(f))
    colmap = json.loads(Path(args.backend).with_suffix(".colmap.json").read_text())
    header = backend_rows[colmap["_header_row_idx"]]
    data_start = colmap["_data_starts_at"]
    data = backend_rows[data_start:]

    # Parse row range
    if args.rows:
        lo, hi = args.rows.split("-")
        lo, hi = int(lo), int(hi)
    else:
        # Infer from citations
        row_ids = [int(c["row_id"]) for c in citations.get("cells", [])
                   if c["row_id"].isdigit()]
        lo, hi = (min(row_ids), max(row_ids)) if row_ids else (1, len(data))

    # Filter to batch rows. row_id is at column colmap["row_id"].
    ci_row = colmap["row_id"]
    batch = [r for r in data if ci_row is not None and len(r) > ci_row
             and r[ci_row].isdigit() and lo <= int(r[ci_row]) <= hi]

    wb = Workbook()

    # README
    build_readme(wb, citations.get("batch_label", f"[ref]-fill batch rows {lo}-{hi}"), [
        f"Batch rows: {lo}-{hi}  ({len(batch)} rows in scope)",
        f"Citation cells: {len(citations.get('cells', []))}",
        f"QA log entries: {len(citations.get('qa_log', []))}",
        f"Data conflicts flagged: {len(citations.get('data_conflicts', []))}",
        f"Candidate data fills: {len(citations.get('candidate_data_fills', []))}",
        f"Defects corrected: {len(citations.get('defects_corrected', []))}",
        "",
        "Color coding:",
        "  Green  = high confidence (multi-source or primary/regulatory + value verbatim)",
        "  Yellow = medium confidence (entity-level or contested)",
        "  Red    = low confidence / review needed",
        "  Peach  = override of pre-existing backend [ref]",
        "  Gray   = pre-existing backend value, untouched",
        "",
        "See QA_review sheet for per-cell provenance log.",
    ])

    # backend_ref_fill sheet
    ws = wb.create_sheet("backend_ref_fill")
    for col_i, h in enumerate(header, start=1):
        ws.cell(row=1, column=col_i, value=h)
    _apply_header_style(ws, 1)

    # Build a citation lookup: (row_id, field) -> citation cell
    cite_lookup = {}
    for c in citations.get("cells", []):
        key = (str(c["row_id"]), c["field"])
        cite_lookup.setdefault(key, []).append(c)

    # Render batch rows
    for r_offset, row in enumerate(batch, start=2):
        row_id = row[ci_row] if len(row) > ci_row else ""
        for col_i, val in enumerate(row, start=1):
            cell = ws.cell(row=r_offset, column=col_i, value=val)
            cell.alignment = WRAP_ALIGN
            # Check if this is a citation cell
            # Find canonical name for this column
            field_name = None
            for fk, fv in colmap.items():
                if fv == col_i - 1 and not fk.startswith("_"):
                    field_name = fk
                    break
            cites = cite_lookup.get((row_id, field_name), [])
            if cites:
                # New citation — apply confidence color and combine URLs.
                # Multiple URLs in one cell join with ", " (SOP §4.15), never
                # newlines.
                conf = cites[0]["confidence"]
                fill = CONFIDENCE_FILLS.get(conf, FILL_RED)
                cell.fill = fill
                cell.value = ", ".join(c["url"] for c in cites)
            elif val and field_name and field_name.endswith("_ref"):
                # Pre-existing ref URL — gray
                cell.fill = FILL_GRAY

    # Freeze + widths
    ws.freeze_panes = "C2"
    for col_i, h in enumerate(header, start=1):
        col_letter = get_column_letter(col_i)
        if "[ref]" in h:
            ws.column_dimensions[col_letter].width = 50
        elif h.lower() in ("name", "shipowner", "shipbuilder", "operator/charterer"):
            ws.column_dimensions[col_letter].width = 25
        else:
            ws.column_dimensions[col_letter].width = 15

    # QA_review sheet
    ws_qa = wb.create_sheet("QA_review")
    qa_row = 1
    sections = [
        ("Per-cell citation log", ["row_id", "field", "action", "confidence", "url", "note"], citations.get("qa_log", [])),
        ("Data-value conflicts", ["row_id", "field", "backend_value", "research_value", "source_url", "note"], citations.get("data_conflicts", [])),
        ("Candidate data-value fills", ["row_id", "field", "proposed_value", "source_url", "confidence", "note"], citations.get("candidate_data_fills", [])),
        ("Defects corrected", ["row_id", "field", "old_url", "new_url", "reason"], citations.get("defects_corrected", [])),
        ("URL verification log", ["url", "status", "soft_error", "content_match", "result"], citations.get("verification_log", [])),
    ]
    for title, cols, items in sections:
        if not items:
            continue
        ws_qa.cell(row=qa_row, column=1, value=title).font = Font(bold=True, size=12)
        qa_row += 1
        for col_i, c in enumerate(cols, start=1):
            cell = ws_qa.cell(row=qa_row, column=col_i, value=c)
            cell.font = HEADER_FONT
            cell.fill = FILL_HEADER
            cell.alignment = HEADER_ALIGN
        qa_row += 1
        for item in items:
            for col_i, c in enumerate(cols, start=1):
                ws_qa.cell(row=qa_row, column=col_i,
                           value=str(item.get(c, ""))).alignment = WRAP_ALIGN
            qa_row += 1
        qa_row += 1  # blank row between sections

    # widths
    for col_letter in "ABCDEF":
        ws_qa.column_dimensions[col_letter].width = 25
    ws_qa.column_dimensions["E"].width = 60  # url column tends to be wide

    out_path = _resolve_out_path(args.out, "lng_carrier_backend_ref_fill.xlsx")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"  Wrote {out_path}")
    return out_path


# Discovery yard-location autofill (Discovery SOP §6.7). The yard-location block
# is a property of the shipbuilder (yard), not the individual vessel, so on a
# candidate row these 7 columns are filled from the authoritative
# data/shipbuilder_facts.csv table (with the backend-sibling scan as a
# fallback). If the shipbuilder is new to both, they stay blank. row_data must
# NOT carry these columns — this autofill owns them. YARD_LOCATION_COLS is the
# canonical list in lookups.YARD_FACT_COLS (imported above).


def _build_yard_location_map(backend_data, backend_header):
    """tag -> {yard-location header: value} for autofill (Discovery SOP §6.7).

    For each normalized shipbuilder tag, picks the existing backend row with the
    most fully-populated yard-location block and records its 7 values. A yard's
    location is constant across its vessels, so any populated row is
    authoritative; most-populated wins so the copied block is complete.
    """
    hidx = {h: i for i, h in enumerate(backend_header)}
    if "Shipbuilder" not in hidx:
        return {}
    sb_i = hidx["Shipbuilder"]
    yl_idx = [(h, hidx[h]) for h in YARD_LOCATION_COLS if h in hidx]
    best = {}  # tag -> (populated_count, block)
    for r in backend_data:
        if len(r) <= sb_i or not r[sb_i].strip():
            continue
        tag = normalize_builder(r[sb_i])
        if not tag:
            continue
        block = {h: (r[i] if len(r) > i else "") for h, i in yl_idx}
        score = sum(1 for v in block.values() if v.strip())
        if score and (tag not in best or score > best[tag][0]):
            best[tag] = (score, block)
    return {tag: block for tag, (_, block) in best.items()}


def _yard_location_map_table_first(backend_data, backend_header):
    """Yard-location map with the authoritative data table taking priority.

    Starts from the backend-sibling scan (_build_yard_location_map) and overlays
    data/shipbuilder_facts.csv: a non-blank table cell wins; cells the table
    leaves blank fall back to the sibling value. A builder absent from the table
    is fully sibling-derived, so this never regresses prior coverage.
    """
    merged = _build_yard_location_map(backend_data, backend_header)
    for tag, block in load_builder_facts().items():
        usable = {k: v for k, v in block.items()
                  if k in YARD_LOCATION_COLS and v and v != AMBIGUOUS}
        if not usable:
            continue
        row = dict(merged.get(tag, {}))
        row.update(usable)
        merged[tag] = row
    return merged


def build_discovery(args):
    """Build the discovery mode workbook."""
    candidates_path = Path(args.candidates)
    payload = json.loads(candidates_path.read_text())

    wb = Workbook()

    build_readme(wb, f"LNG Carrier Discovery — gap window {payload.get('gap_window', 'unspecified')}", [
        f"Yards swept: {len(payload.get('yards_swept', []))}",
        f"Candidate clusters: {len(payload.get('candidates', []))}",
        f"Backend status flags: {len(payload.get('backend_status_flags', []))}",
        "",
        "Color coding (on candidate_vessels sheet):",
        "  Green  = 2+ cross-checked sources OR 1 primary/regulatory source with value verbatim",
        "  Yellow = entity-level confirmation; some data point implied/contested",
        "  Red    = single source / weak corroboration — review before promoting",
        "",
        "Workflow: review candidates -> approve/reject -> manually paste approved rows into backend.",
    ])

    # candidate_vessels sheet — mirrors the LIVE backend column order EXACTLY
    # after the four prefix columns (Discovery SOP §5.2 / §6.6 paste-compat).
    # The header is read from the fresh backend pull rather than hardcoded, so
    # the sheet always tracks the current schema (geolocation, Researcher, Last
    # updated, [Original source], etc. all appear as blank columns).
    with open(args.backend, encoding="utf-8") as f:
        backend_rows = list(csv.reader(f))
    colmap = json.loads(Path(args.backend).with_suffix(".colmap.json").read_text())
    backend_header = backend_rows[colmap["_header_row_idx"]]
    header_index = {h: i for i, h in enumerate(backend_header)}
    data_start = colmap.get("_data_starts_at", colmap["_header_row_idx"] + 1)
    yard_loc_map = _yard_location_map_table_first(backend_rows[data_start:], backend_header)

    ws = wb.create_sheet("candidate_vessels")
    prefix_cols = ["cluster_id", "cluster_label", "confidence", "discovery_notes"]
    headers = prefix_cols + backend_header
    for col_i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_i, value=h)
    _apply_header_style(ws, 1)

    n_prefix = len(prefix_cols)
    for r_offset, cand in enumerate(payload.get("candidates", []), start=2):
        confidence = cand.get("confidence", "Y")
        fill = CONFIDENCE_FILLS.get(confidence, FILL_YELLOW)
        # prefix cols
        ws.cell(row=r_offset, column=1, value=cand.get("cluster_id", "")).fill = fill
        ws.cell(row=r_offset, column=2, value=cand.get("cluster_label", "")).fill = fill
        ws.cell(row=r_offset, column=3, value=confidence).fill = fill
        ws.cell(row=r_offset, column=4, value=cand.get("discovery_notes", "")).fill = fill
        # backend cols — row_data is keyed by EXACT backend header strings.
        row_data = dict(cand.get("row_data", {}))
        # Yard-location autofill (Discovery SOP §6.7): copy the 7 yard-location
        # columns from an existing backend row for the same shipbuilder; leave
        # blank if the shipbuilder is new. This autofill owns those columns —
        # any values for them in row_data are dropped.
        provided_yl = [k for k in YARD_LOCATION_COLS if k in row_data]
        for k in YARD_LOCATION_COLS:
            row_data.pop(k, None)
        sb_name = row_data.get("Shipbuilder", "")
        tag = normalize_builder(sb_name) if sb_name else ""
        if tag and tag in yard_loc_map:
            row_data.update(yard_loc_map[tag])
            if provided_yl:
                print(f"  [info] candidate {cand.get('cluster_id')}: yard-location "
                      f"columns {provided_yl} autofilled from backend (shipbuilder "
                      f"{sb_name!r}); row_data values ignored (SOP §6.7).",
                      file=sys.stderr)
        elif provided_yl:
            print(f"  [info] candidate {cand.get('cluster_id')}: shipbuilder "
                  f"{sb_name!r} not in backend; yard-location columns left blank "
                  f"(SOP §6.7).", file=sys.stderr)
        unknown = [k for k in row_data if k not in header_index]
        if unknown:
            print(f"  [warn] candidate {cand.get('cluster_id')}: row_data keys not in "
                  f"backend header (ignored): {unknown}", file=sys.stderr)
        for h, idx in header_index.items():
            val = row_data.get(h, "")
            if val and h.endswith("[ref]") and "\n" in val:
                # Multiple URLs in one [ref] cell join with ", " (SOP §4.15).
                val = ", ".join(p.strip() for p in val.split("\n") if p.strip())
            cell = ws.cell(row=r_offset, column=n_prefix + 1 + idx, value=val)
            cell.alignment = WRAP_ALIGN
            if val:  # only color cells that are populated
                cell.fill = fill

    ws.freeze_panes = "E2"
    for col_i, h in enumerate(headers, start=1):
        col_letter = get_column_letter(col_i)
        if "[ref]" in h:
            ws.column_dimensions[col_letter].width = 50
        elif h in ("cluster_label", "discovery_notes", "Name",
                   "Shipowner", "Shipbuilder", "Operator/charterer"):
            ws.column_dimensions[col_letter].width = 25
        else:
            ws.column_dimensions[col_letter].width = 15

    # QA_review sheet
    ws_qa = wb.create_sheet("QA_review")
    qa_row = 1
    sections = [
        ("Per-candidate provenance log",
         ["cluster_id", "cluster_label", "confidence", "source_urls", "notes"],
         payload.get("provenance_log", [])),
        ("Backend status flags",
         ["row_id", "issue_type", "details", "suggested_action"],
         payload.get("backend_status_flags", [])),
        ("URL verification log",
         ["url", "status", "soft_error", "content_match", "result"],
         payload.get("verification_log", [])),
        ("Search methodology audit",
         ["yard", "rows_in_window", "candidates_found", "notes"],
         payload.get("methodology_audit", {}).get("yards", [])),
    ]
    for title, cols, items in sections:
        if not items:
            continue
        ws_qa.cell(row=qa_row, column=1, value=title).font = Font(bold=True, size=12)
        qa_row += 1
        for col_i, c in enumerate(cols, start=1):
            cell = ws_qa.cell(row=qa_row, column=col_i, value=c)
            cell.font = HEADER_FONT
            cell.fill = FILL_HEADER
        qa_row += 1
        for item in items:
            for col_i, c in enumerate(cols, start=1):
                ws_qa.cell(row=qa_row, column=col_i,
                           value=str(item.get(c, ""))).alignment = WRAP_ALIGN
            qa_row += 1
        qa_row += 1

    for col_letter in "ABCDE":
        ws_qa.column_dimensions[col_letter].width = 30

    # backend_status_flags sheet (mirror of QA section 2)
    if payload.get("backend_status_flags"):
        ws_bsf = wb.create_sheet("backend_status_flags")
        cols = ["row_id", "issue_type", "details", "suggested_action"]
        for col_i, c in enumerate(cols, start=1):
            ws_bsf.cell(row=1, column=col_i, value=c)
        _apply_header_style(ws_bsf, 1)
        for r_offset, item in enumerate(payload["backend_status_flags"], start=2):
            for col_i, c in enumerate(cols, start=1):
                ws_bsf.cell(row=r_offset, column=col_i,
                            value=str(item.get(c, ""))).alignment = WRAP_ALIGN
        for col_letter in "ABCD":
            ws_bsf.column_dimensions[col_letter].width = 30

    out_path = _resolve_out_path(args.out, "lng_carrier_candidate_vessels.xlsx")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"  Wrote {out_path}")
    return out_path


def _join_refs(existing: str, new_urls) -> str:
    """Combine an existing [ref] cell value with new corroborator URLs.

    The existing URL(s) come FIRST and are never dropped (the data-fill preserve
    rule, Data-fill SOP §4). All URLs are de-duplicated and joined with ", "
    (RF §4.15); any newline in the existing value is normalized to ", " too.
    """
    out = []

    def _add(chunk):
        for part in str(chunk).replace("\n", ", ").split(", "):
            part = part.strip()
            if part and part not in out:
                out.append(part)

    _add(existing or "")
    for u in (new_urls or []):
        _add(u)
    return ", ".join(out)


# Controlled vocabularies for the type columns (Data-fill SOP §8). The backend
# writes values verbatim with no normalizer, so a data_fill proposal MUST use an
# existing canonical value. Single source of truth: lookups.CONTROLLED_VOCAB
# (mirrors data/controlled_vocab.md), shared with qc_backend.py.
_DATA_FILL_VOCAB = CONTROLLED_VOCAB


def _validate_data_fills(fills, header_index, row_by_id):
    """Print non-fatal Data-fill consistency warnings: orphan refs, Price without
    currency, Capacity without units, off-vocab type values."""
    issues = 0

    def _has(rid, header):
        row = row_by_id.get(rid, [])
        idx = header_index.get(header)
        return idx is not None and len(row) > idx and bool(row[idx].strip())

    by_row = {}
    for f in fills:
        by_row.setdefault(str(f["row_id"]), {})[f.get("field", "")] = f
    for rid, fmap in by_row.items():
        if "Price" in fmap and "Price currency" not in fmap and not _has(rid, "Price currency"):
            print(f"  [warn] row {rid}: Price proposed without a Price currency", file=sys.stderr); issues += 1
        if "Capacity" in fmap and "Capacity units" not in fmap and not _has(rid, "Capacity units"):
            print(f"  [warn] row {rid}: Capacity proposed without Capacity units", file=sys.stderr); issues += 1
    for f in fills:
        rid, fld = str(f["row_id"]), f.get("field", "")
        if fld in _DATA_FILL_VOCAB and f.get("proposed_value", "") not in _DATA_FILL_VOCAB[fld]:
            print(f"  [warn] row {rid}: {fld} value {f.get('proposed_value')!r} not in controlled vocab",
                  file=sys.stderr); issues += 1
        if f.get("new_urls") and not f.get("proposed_value") and not _has(rid, fld):
            print(f"  [warn] row {rid}: {fld} has URL(s) but no value (orphan-ref risk)",
                  file=sys.stderr); issues += 1
    if issues:
        print(f"  [warn] data_fill: {issues} consistency warning(s) — review before promoting",
              file=sys.stderr)
    return issues


def build_data_fill(args):
    """Build the data-fill mode workbook (Data-fill SOP).

    Proposes values + corroborating [ref]s for BLANK and literal-`unknown`
    backend data cells, for human review. NEVER edits the backend (RF §4.7).
    `unknown` cells are treated as blank for research, but their existing [ref]
    URL(s) are preserved and only appended to (DF §4).
    """
    payload = json.loads(Path(args.fills).read_text())

    with open(args.backend, encoding="utf-8") as f:
        backend_rows = list(csv.reader(f))
    colmap = json.loads(Path(args.backend).with_suffix(".colmap.json").read_text())
    backend_header = backend_rows[colmap["_header_row_idx"]]
    header_index = {h: i for i, h in enumerate(backend_header)}
    data_start = colmap.get("_data_starts_at", colmap["_header_row_idx"] + 1)
    ci_row = colmap["row_id"]
    row_by_id = {r[ci_row].strip(): r for r in backend_rows[data_start:]
                 if len(r) > ci_row and r[ci_row].strip()}

    fills = payload.get("fills", [])
    scope_ids = [str(x) for x in payload.get("scope", {}).get("row_ids", [])]
    if not scope_ids:
        scope_ids = sorted({str(f["row_id"]) for f in fills},
                           key=lambda s: int(s) if s.isdigit() else 1 << 30)

    CONF_RANK = {"G": 3, "green": 3, "Y": 2, "yellow": 2, "R": 1, "red": 1}
    # corroborate fills leave the data value untouched (they only append [ref]s), so
    # they are excluded from data_fill_by — the value cell stays gray, not painted as
    # a proposed change; they still flow through ref_fill_by to get the peach append.
    data_fill_by = {(str(f["row_id"]), f["field"]): f for f in fills
                    if f.get("field") and f.get("prev_state") != "corroborate"}
    ref_fill_by = {(str(f["row_id"]), f["ref_field"]): f for f in fills if f.get("ref_field")}

    wb = Workbook()
    n_derivable = sum(1 for f in fills if f.get("derivable"))
    build_readme(wb, payload.get("batch_label", "LNG Carrier Data-fill"), [
        f"In-scope rows: {len(scope_ids)}",
        f"Proposed fills: {len(fills)}  (derivable autofill: {n_derivable}, "
        f"researched: {len(fills) - n_derivable})",
        f"Documented blanks (researched, not found): {len(payload.get('documented_blanks', []))}",
        "",
        "Color coding (backend_data_fill sheet):",
        "  Gray   = pre-existing backend value, untouched",
        "  Green/Yellow/Red = PROPOSED fill (confidence) for a blank or 'unknown' cell",
        "  Peach  = [ref] cell that already had URL(s): existing URL kept FIRST, "
        "corroborator appended (never replaced)",
        "",
        "'unknown' data cells are treated as blank for research; the proposed value carries a",
        "cell comment 'prev: unknown' and the existing [ref] URL(s) are preserved (peach).",
        "",
        "Workflow: review proposals -> accept/reject -> manually enter accepted value + [ref] "
        "into the backend.",
        "The backend is NEVER edited by this tool (RF §4.7).",
    ])

    ws = wb.create_sheet("backend_data_fill")
    prefix_cols = ["row_id", "cluster_id", "cluster_label", "n_proposed", "max_confidence"]
    n_prefix = len(prefix_cols)
    headers = prefix_cols + backend_header
    for col_i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_i, value=h)
    _apply_header_style(ws, 1)

    missing = []
    r_offset = 2
    for rid in scope_ids:
        row = row_by_id.get(rid)
        if row is None:
            missing.append(rid)
            continue

        def at(idx):
            return row[idx] if idx is not None and len(row) > idx else ""

        cluster_label = f"{normalize_builder(at(header_index.get('Shipbuilder')))}|" \
                        f"{normalize_owner(at(header_index.get('Shipowner')))}"
        row_fills = [f for f in fills if str(f["row_id"]) == rid]
        max_rank = max((CONF_RANK.get(f.get("confidence", "R"), 1) for f in row_fills), default=0)
        ws.cell(row=r_offset, column=1, value=rid)
        ws.cell(row=r_offset, column=2, value=cluster_label)
        ws.cell(row=r_offset, column=3, value=cluster_label)
        ws.cell(row=r_offset, column=4, value=len(row_fills))
        ws.cell(row=r_offset, column=5, value={3: "G", 2: "Y", 1: "R", 0: ""}[max_rank])

        for h, idx in header_index.items():
            col = n_prefix + 1 + idx
            existing = at(idx)
            if (rid, h) in data_fill_by:                      # proposed data value
                e = data_fill_by[(rid, h)]
                c = ws.cell(row=r_offset, column=col, value=e.get("proposed_value", ""))
                c.fill = CONFIDENCE_FILLS.get(e.get("confidence", "R"), FILL_RED)
                if e.get("prev_state") == "unknown":
                    c.comment = Comment('prev: "unknown" (researched & proposed)', "data_fill")
            elif (rid, h) in ref_fill_by:                     # [ref] cell getting corroborator(s)
                e = ref_fill_by[(rid, h)]
                new_urls = e.get("new_urls", [])
                e["existing_ref_preserved"] = existing        # from FRESH backend, for the QA log
                c = ws.cell(row=r_offset, column=col, value=_join_refs(existing, new_urls))
                if existing and new_urls:
                    c.fill = FILL_PEACH
                elif new_urls:
                    c.fill = CONFIDENCE_FILLS.get(e.get("confidence", "R"), FILL_RED)
                elif existing:
                    c.fill = FILL_GRAY
            else:                                             # untouched pre-existing value
                if existing and h.endswith("[ref]") and "\n" in existing:
                    existing = ", ".join(p.strip() for p in existing.split("\n") if p.strip())
                c = ws.cell(row=r_offset, column=col, value=existing)
                if existing:
                    c.fill = FILL_GRAY
            c.alignment = WRAP_ALIGN
        r_offset += 1

    if missing:
        print(f"  [warn] scope row_ids not found in backend (skipped): {missing}", file=sys.stderr)
    _validate_data_fills(fills, header_index, row_by_id)

    ws.freeze_panes = "F2"
    for col_i, h in enumerate(headers, start=1):
        letter = get_column_letter(col_i)
        if "[ref]" in h:
            ws.column_dimensions[letter].width = 50
        elif h in ("cluster_id", "cluster_label", "Name", "Shipowner",
                   "Shipbuilder", "Operator/charterer"):
            ws.column_dimensions[letter].width = 24
        else:
            ws.column_dimensions[letter].width = 15

    # QA_review
    ws_qa = wb.create_sheet("QA_review")
    qa_fills = []
    for f in fills:
        g = dict(f)
        nu = g.get("new_urls", [])
        g["new_urls"] = ", ".join(nu) if isinstance(nu, list) else str(nu)
        # Default accept/hold mirrors apply_batch.py / decisions.csv (which is the
        # authoritative decision surface). Green/derivable -> accept, else hold.
        g["decision"] = "accept" if (g.get("derivable") or g.get("confidence") == "G") else "hold"
        qa_fills.append(g)
    qa_row = 1
    sections = [
        ("Candidate data-value fills",
         ["row_id", "field", "prev_state", "proposed_value", "new_urls",
          "existing_ref_preserved", "confidence", "decision", "note"], qa_fills),
        ("Documented blanks (researched, not found)",
         ["row_id", "field", "searched", "as_of", "note"],
         payload.get("documented_blanks", [])),
        ("URL verification log",
         ["url", "status", "soft_error", "content_match", "result"],
         payload.get("verification_log", [])),
    ]
    for title, cols, items in sections:
        if not items:
            continue
        ws_qa.cell(row=qa_row, column=1, value=title).font = Font(bold=True, size=12)
        qa_row += 1
        for col_i, c in enumerate(cols, start=1):
            cell = ws_qa.cell(row=qa_row, column=col_i, value=c)
            cell.font = HEADER_FONT
            cell.fill = FILL_HEADER
        qa_row += 1
        for item in items:
            for col_i, c in enumerate(cols, start=1):
                ws_qa.cell(row=qa_row, column=col_i,
                           value=str(item.get(c, ""))).alignment = WRAP_ALIGN
            qa_row += 1
        qa_row += 1
    for letter in "ABCDEFGH":
        ws_qa.column_dimensions[letter].width = 22
    ws_qa.column_dimensions["E"].width = 60

    out_path = _resolve_out_path(args.out, "lng_carrier_data_fill.xlsx")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"  Wrote {out_path}")
    return out_path


def build_fix(args):
    """Build a correction-batch workbook (column-offset / value / ref repairs).

    Unlike the earlier hand-built CSV corrections (which bypassed every gate and
    is how the 176,400-vs-180,000 capacity defect shipped), a fix batch routes
    every corrected (value, [ref]) pair through the §3.8 value↔ref corroboration
    gate: a ref may only stay on a cell whose value its live page actually
    contains. Non-corroborating refs are dropped and surfaced in QA_review;
    explicitly listed drop_refs are removed; soft-blocked refs (HTTP 000/403,
    e.g. Cloudflare) are kept ONLY when flagged soft:true (a human attests §3.8a
    out-of-band content confirmation) and are commented as such.

    fix.json schema:
      {
        "batch_label": "Fix — rows 1216/1217 capacity",
        "reason": "free-text why",
        "corrections": [
          { "row_id": "1216",
            "cells": [
              { "field": "Capacity",            # backend value column header
                "new_value": "180000",
                "confidence": "G",
                "refs": [ {"url": "https://...", "soft": false}, ... ],
                "drop_refs": ["http://...csb..."],   # remove from the [ref] cell
                "note": "CSB 176,400 = 98% figure; nominal 180,000 per 3 sources" }
            ] } ] }

    preserve_ref:true on a cell is the escape hatch for cosmetic / derived-value
    edits (e.g. normalizing a synthetic placeholder Name like
    "Hudong-Zhonghua Shanghai (MISC 1)" -> "Hudong-Zhonghua (MISC 1)"). The paired
    [ref] there documents the underlying ORDER (builder/owner/contract), not the
    literal name string, so the §3.8c value↔ref gate doesn't apply — running it
    would wrongly drop good discovery refs (no page contains the exact placeholder
    string). With preserve_ref:true the value is rewritten, the existing [ref] cell
    is left byte-for-byte untouched, the gate is skipped, and a PRESERVED line is
    logged to QA_review. `refs`/`drop_refs` are ignored on such a cell.

    Each cell sets both the value column (`field`) and its paired `field + ' [ref]'`
    column. The corrected rows are emitted full-width in backend column order, so
    they paste straight over the matching backend rows. The backend is NEVER
    edited here (RF §4.7).
    """
    from url_verifier import corroborates

    payload = json.loads(Path(args.fix).read_text())
    with open(args.backend, encoding="utf-8") as f:
        backend_rows = list(csv.reader(f))
    colmap = json.loads(Path(args.backend).with_suffix(".colmap.json").read_text())
    backend_header = backend_rows[colmap["_header_row_idx"]]
    header_index = {h: i for i, h in enumerate(backend_header)}
    data_start = colmap.get("_data_starts_at", colmap["_header_row_idx"] + 1)
    ci_row = colmap["row_id"]
    row_by_id = {r[ci_row].strip(): r for r in backend_rows[data_start:]
                 if len(r) > ci_row and r[ci_row].strip()}

    # --base: build on an already-reviewed corrected-rows CSV instead of the live
    # backend row. Use this when the live row is still corrupted (e.g. a prior
    # structural fix wasn't applied yet) — the base supplies the correct full-row
    # layout, this batch's cells apply the gated value/ref corrections on top.
    # Rows are remapped to live-backend column order BY HEADER NAME, so any
    # trailing/stray columns absent from the base are cleared.
    if getattr(args, "base", None):
        with open(args.base, encoding="utf-8") as f:
            base_rows = list(csv.reader(f))
        base_header = base_rows[0]
        for br in base_rows[1:]:
            if not any(c.strip() for c in br):
                continue
            d = {base_header[i]: (br[i] if i < len(br) else "")
                 for i in range(len(base_header))}
            rid = d.get(backend_header[ci_row], "").strip()
            if not rid:
                continue
            row_by_id[rid] = [d.get(h, "") if h else "" for h in backend_header]

    corrections = payload.get("corrections", [])
    scope_ids = [str(c["row_id"]) for c in corrections]
    # changed-cell map: (row_id, header) -> {kind, conf, value} for styling
    changed = {}
    qa_rows = []          # per-ref verification results for QA_review
    missing = []

    for corr in corrections:
        rid = str(corr["row_id"])
        if rid not in row_by_id:
            missing.append(rid)
            continue
        for cell in corr.get("cells", []):
            field = cell["field"]
            ref_field = field + " [ref]"
            new_value = str(cell.get("new_value", ""))
            existing_ref = ""
            if ref_field in header_index:
                ri = header_index[ref_field]
                cur = row_by_id[rid]
                existing_ref = cur[ri] if len(cur) > ri else ""
            preserve_ref = bool(cell.get("preserve_ref"))
            drop_set = set() if preserve_ref else set(cell.get("drop_refs", []))

            kept = []
            if preserve_ref:
                qa_rows.append({"row_id": rid, "field": field, "value": new_value,
                                "url": existing_ref,
                                "verdict": "PRESERVED (cosmetic/derived value; paired [ref] "
                                           "documents the order, not the literal string — "
                                           "§3.8c gate N/A)",
                                "note": cell.get("note", "")})
            for ref in (cell.get("refs", []) if not preserve_ref else []):
                url = ref["url"] if isinstance(ref, dict) else ref
                soft = bool(ref.get("soft")) if isinstance(ref, dict) else False
                ok, reason = corroborates(url, new_value)
                if ok:
                    kept.append(url)
                    verdict = "PASS (corroborates)"
                elif soft and (reason.startswith("HTTP") or "soft-error" in reason):
                    kept.append(url)
                    verdict = f"SOFT-KEPT §3.8a ({reason}; human-confirmed off-band)"
                elif reason.startswith("HTTP") or "soft-error" in reason:
                    verdict = f"DROPPED — unreachable, not §3.8a-flagged ({reason})"
                else:
                    verdict = f"DROPPED — does NOT corroborate {new_value!r} ({reason})"
                qa_rows.append({"row_id": rid, "field": field, "value": new_value,
                                "url": url, "verdict": verdict, "note": cell.get("note", "")})
            for u in drop_set:
                qa_rows.append({"row_id": rid, "field": field, "value": new_value,
                                "url": u, "verdict": "REMOVED (drop_refs: conflicts with value)",
                                "note": cell.get("note", "")})

            # write into the working backend row copy
            cur = list(row_by_id[rid])
            vi = header_index.get(field)
            if vi is not None:
                while len(cur) <= vi:
                    cur.append("")
                cur[vi] = new_value
                changed[(rid, field)] = {"kind": "value", "conf": cell.get("confidence", "G")}
            if ref_field in header_index and not preserve_ref:
                ri = header_index[ref_field]
                while len(cur) <= ri:
                    cur.append("")
                joined = ", ".join(kept)
                cur[ri] = joined
                changed[(rid, ref_field)] = {"kind": "ref", "had_existing": bool(existing_ref)}
            row_by_id[rid] = cur

    if missing:
        print(f"  [warn] correction row_ids not in backend (skipped): {missing}", file=sys.stderr)

    dropped = [q for q in qa_rows if q["verdict"].startswith(("DROPPED", "REMOVED"))]

    wb = Workbook()
    build_readme(wb, payload.get("batch_label", "LNG Carrier Fix"), [
        payload.get("reason", ""),
        "",
        f"Rows corrected: {len(scope_ids)}   Refs checked: {len(qa_rows)}   "
        f"Refs dropped/removed: {len(dropped)}",
        "",
        "Every corrected (value, [ref]) pair passed the §3.8 value↔ref corroboration",
        "gate: a ref stays only if its live page contains the cell's value. Refs that",
        "named a DIFFERENT value were dropped (see QA_review). The backend is NEVER",
        "edited by this tool (RF §4.7) — paste the corrected rows over the matching",
        "backend rows.",
        "",
        "Color coding (fix sheet):",
        "  Green/Yellow/Red = corrected value (confidence)",
        "  Peach = [ref] cell rewritten (corroborating refs only)",
        "  Gray  = pre-existing cell, untouched",
    ])

    ws = wb.create_sheet("fix")
    for col_i, h in enumerate(backend_header, start=1):
        ws.cell(row=1, column=col_i, value=h)
    _apply_header_style(ws, 1)
    r_offset = 2
    for rid in scope_ids:
        row = row_by_id.get(rid)
        if row is None:
            continue
        for idx, h in enumerate(backend_header):
            val = row[idx] if len(row) > idx else ""
            c = ws.cell(row=r_offset, column=idx + 1, value=val)
            ch = changed.get((rid, h))
            if ch and ch["kind"] == "value":
                c.fill = CONFIDENCE_FILLS.get(ch["conf"], FILL_GREEN)
            elif ch and ch["kind"] == "ref":
                c.fill = FILL_PEACH
            elif val:
                c.fill = FILL_GRAY
            c.alignment = WRAP_ALIGN
        r_offset += 1
    ws.freeze_panes = "B2"
    for col_i, h in enumerate(backend_header, start=1):
        letter = get_column_letter(col_i)
        ws.column_dimensions[letter].width = 50 if "[ref]" in h else 18

    ws_qa = wb.create_sheet("QA_review")
    cols = ["row_id", "field", "value", "url", "verdict", "note"]
    for col_i, c in enumerate(cols, start=1):
        cell = ws_qa.cell(row=1, column=col_i, value=c)
        cell.font = HEADER_FONT
        cell.fill = FILL_HEADER
    for r_i, q in enumerate(qa_rows, start=2):
        for col_i, c in enumerate(cols, start=1):
            ws_qa.cell(row=r_i, column=col_i, value=str(q.get(c, ""))).alignment = WRAP_ALIGN
    for letter, w in {"A": 8, "B": 16, "C": 12, "D": 60, "E": 44, "F": 50}.items():
        ws_qa.column_dimensions[letter].width = w
    ws_qa.freeze_panes = "A2"

    out_path = _resolve_out_path(args.out, "lng_carrier_fix.xlsx")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"  Wrote {out_path}  ({len(dropped)} ref(s) dropped by the gate)")
    return out_path


# --- FSRU reconciliation mode (GIIGNL fleet table ↔ backend) -------------------
def _table_sheet(wb, title, cols, rows, fills=None, widths=None, header_note=None):
    """Render a simple header+rows sheet. `rows` is a list of dicts keyed by the
    column names in `cols`. `fills` is an optional list (parallel to rows) of a
    PatternFill or None to color that data row. `widths` maps column-letter->width.
    `header_note` is an italic line written above the table."""
    ws = wb.create_sheet(title)
    top = 1
    if header_note:
        c = ws.cell(row=1, column=1, value=header_note)
        c.font = Font(name="Calibri", size=10, italic=True)
        c.alignment = WRAP_ALIGN
        top = 2
    for col_i, h in enumerate(cols, start=1):
        ws.cell(row=top, column=col_i, value=h)
    _apply_header_style(ws, top)
    for r_off, item in enumerate(rows):
        fill = (fills[r_off] if fills else None)
        for col_i, c in enumerate(cols, start=1):
            cell = ws.cell(row=top + 1 + r_off, column=col_i, value=item.get(c, ""))
            cell.alignment = WRAP_ALIGN
            if fill:
                cell.fill = fill
    if widths:
        for letter, w in widths.items():
            ws.column_dimensions[letter].width = w
    ws.freeze_panes = ws.cell(row=top + 1, column=1).coordinate
    return ws


def _yn(v):
    """Tri-state agree flag -> human label."""
    return {True: "yes", False: "NO", None: "?"}.get(v, "?")


def build_fsru(args):
    """Build the FSRU reconciliation workbook from work/fsru_reconcile.json.

    Renders the GIIGNL↔backend buckets for human review. GIIGNL is a COMPARISON
    ARTIFACT, not a citable [ref] (same status as SFOC): GIIGNL-derived values are
    starting points to verify, and NO [ref] cells are pre-filled — a primary
    source must be attached via url_verifier.py before any candidate is promoted.
    The backend is never edited here (additive candidate workbook only)."""
    rec = json.loads(Path(args.reconcile).read_text())
    s = rec["summary"]

    with open(args.backend, encoding="utf-8") as f:
        backend_rows = list(csv.reader(f))
    colmap = json.loads(Path(args.backend).with_suffix(".colmap.json").read_text())
    backend_header = backend_rows[colmap["_header_row_idx"]]
    header_index = {h: i for i, h in enumerate(backend_header)}

    wb = Workbook()
    build_readme(wb, f"FSRU reconciliation — GIIGNL {rec.get('edition_year')} fleet vs backend", [
        f"GIIGNL in-service FSRU fleet: {rec['fleet_count']}   |   backend FSRUs: {rec['backend_fsru_count']}",
        "",
        "Buckets (one sheet each):",
        f"  Matched_with_diffs   {s['matched']:>3}  GIIGNL FSRU already in backend as FSRU (field diffs flagged)",
        f"  Reclassify           {s['reclassify']:>3}  GIIGNL FSRU present but typed non-FSRU in backend (typing finding)",
        f"  Manual_pairing       {s['manual']:>3}  no name match; capacity+owner suggest a backend row (human pairs)",
        f"  Candidates_to_add    {s['candidates']:>3}  in GIIGNL, absent from backend ({s['candidates_small_scale']} need FSRU-vs-small-scale review)",
        f"  Backend_only         {s['backend_only']:>3}  backend FSRUs absent from GIIGNL (expected: on-order/idle)",
        f"  FSU_exclusions       {s['backend_fsu']:>3}  backend FSUs (storage-only, out of scope) — known exclusions",
        f"  GIIGNL_orderbook     {s['orderbook']:>3}  GIIGNL future deliveries (Phase B reference)",
        "",
        "Join: GIIGNL has no IMO column, so the match is by vessel name over",
        "{current name} ∪ {ex_names}, normalized (diacritics/parentheticals folded).",
        "CAPACITY (storage m³ ↔ backend Capacity cbm) is the corroborator — it agreed",
        "on every matched pair. BUILDER is informational only: for converted units",
        "GIIGNL lists the conversion yard while the backend keeps the original builder.",
        "",
        "Color: Green = full-size gap to add / clean; Yellow = needs a human decision",
        "(small-scale review, reclassification, manual pairing); Gray = expected non-finding.",
        "",
        "SOURCE RULE: GIIGNL is NOT citable (it is downstream of Clarksons; comparison",
        "artifact only). Every promoted value needs a primary [ref] via url_verifier.py.",
        "Promote vetted rows through the apply pipeline — the backend is never auto-edited.",
    ])

    # --- Summary sheet ---------------------------------------------------------
    summary_rows = [
        {"metric": "GIIGNL in-service FSRU fleet", "count": rec["fleet_count"]},
        {"metric": "Backend FSRUs (Vessel type = FSRU)", "count": rec["backend_fsru_count"]},
        {"metric": "— Matched (already FSRU in backend)", "count": s["matched"]},
        {"metric": "— Reclassify (present, typed non-FSRU)", "count": s["reclassify"]},
        {"metric": "— Manual pairing (name defect / no name match)", "count": s["manual"]},
        {"metric": "— Candidates to add (genuine gaps)", "count": s["candidates"]},
        {"metric": "     of which small-scale review", "count": s["candidates_small_scale"]},
        {"metric": "Backend-only FSRUs (absent from GIIGNL)", "count": s["backend_only"]},
        {"metric": "Backend FSUs (excluded, storage-only)", "count": s["backend_fsu"]},
        {"metric": "GIIGNL orderbook (future)", "count": s["orderbook"]},
    ]
    _table_sheet(
        wb, "Summary", ["metric", "count"], summary_rows,
        widths={"A": 48, "B": 10},
        header_note=("Headline: the tracker's coverage of GIIGNL's in-service FSRU fleet is "
                     "essentially complete — every full-size in-service FSRU is present (a few "
                     "mistyped/misnamed); the only genuine absences are small/power-barge units "
                     "pending FSRU-vs-small-scale review."))

    # --- Candidates_to_add (backend column order) ------------------------------
    ws = wb.create_sheet("Candidates_to_add")
    prefix_cols = ["candidate_id", "giignl_name", "small_scale_review", "review_reason",
                   "giignl_owner", "giignl_builder", "giignl_sendout_mtpa",
                   "giignl_terminal", "giignl_status"]
    headers = prefix_cols + backend_header
    for col_i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_i, value=h)
    _apply_header_style(ws, 1)
    n_prefix = len(prefix_cols)
    for r_off, cand in enumerate(rec.get("candidates", []), start=2):
        g = cand["giignl"]
        small = cand["small_scale_review"]
        fill = FILL_YELLOW if small else FILL_GREEN
        prefix = {
            "candidate_id": f"FSRU-C{r_off-1}",
            "giignl_name": g["vessel_name"],
            "small_scale_review": "yes" if small else "no",
            "review_reason": cand["review_reason"],
            "giignl_owner": g.get("vessel_owner", ""),
            "giignl_builder": g.get("builder", ""),
            "giignl_sendout_mtpa": g.get("sendout_mtpa", ""),
            "giignl_terminal": g.get("location_raw", ""),
            "giignl_status": g.get("location_status", ""),
        }
        for col_i, h in enumerate(prefix_cols, start=1):
            ws.cell(row=r_off, column=col_i, value=prefix[h]).fill = fill
        # backend-column starting values: structural fields only. Vessel type is
        # left BLANK for small-scale (that's the open question); [ref] cells stay
        # blank (GIIGNL not citable). Owner/builder live in the prefix to verify.
        row_data = {
            "Name": g["vessel_name"],
            "Capacity": g.get("storage_m3", ""),
            "Capacity units": "cbm" if g.get("storage_m3") else "",
        }
        if not small:
            row_data["Vessel type"] = "FSRU"
        for h, idx in header_index.items():
            val = row_data.get(h, "")
            cell = ws.cell(row=r_off, column=n_prefix + 1 + idx, value=val)
            cell.alignment = WRAP_ALIGN
            if val:
                cell.fill = fill
    ws.freeze_panes = ws.cell(row=2, column=n_prefix + 1).coordinate
    for col_i, h in enumerate(headers, start=1):
        letter = get_column_letter(col_i)
        if "[ref]" in h:
            ws.column_dimensions[letter].width = 40
        elif h in ("review_reason", "giignl_terminal", "Name", "giignl_owner",
                   "giignl_builder", "Shipowner", "Shipbuilder"):
            ws.column_dimensions[letter].width = 26
        else:
            ws.column_dimensions[letter].width = 14

    # --- Matched_with_diffs ----------------------------------------------------
    mcols = ["giignl_name", "backend_name", "live_sheet_row", "imo",
             "cap_giignl", "cap_backend", "cap_agree",
             "built_giignl", "delivery_backend", "delivery_agree",
             "owner_giignl", "owner_backend", "owner_overlap",
             "builder_giignl", "builder_backend", "builder_note"]
    mrows, mfills = [], []
    for r in rec["matched"]:
        g, be, df = r["giignl"], r["backend"], r["diffs"]
        cap_ok = df["capacity"]["agree"]
        yr_ok = df["delivery_year"]["same"]
        own_ok = df["owner"]["overlap"]
        # color only GENUINE diffs: capacity disagreement = red; delivery or owner
        # mismatch = yellow; builder diffs are expected/informational (no color).
        fill = (FILL_RED if cap_ok is False else
                FILL_YELLOW if (yr_ok is False or own_ok is False) else None)
        mrows.append({
            "giignl_name": g["vessel_name"], "backend_name": be["name"],
            "live_sheet_row": be["sheet_row"], "imo": be["imo"],
            "cap_giignl": df["capacity"]["giignl_storage_m3"],
            "cap_backend": df["capacity"]["backend_capacity"], "cap_agree": _yn(cap_ok),
            "built_giignl": df["delivery_year"]["giignl_built"],
            "delivery_backend": df["delivery_year"]["backend_delivery"],
            "delivery_agree": _yn(yr_ok),
            "owner_giignl": df["owner"]["giignl"], "owner_backend": df["owner"]["backend"],
            "owner_overlap": _yn(own_ok),
            "builder_giignl": df["builder"]["giignl"], "builder_backend": df["builder"]["backend"],
            "builder_note": ("conversion-yard/abbrev (informational)"
                             if df["builder"]["same_tag"] is False else ""),
        })
        mfills.append(fill)
    _table_sheet(wb, "Matched_with_diffs", mcols, mrows, fills=mfills,
                 widths={"A": 24, "B": 24, "K": 26, "L": 26, "N": 22, "O": 22, "P": 30},
                 header_note=("Capacity is the join corroborator (agreed on all matches). "
                              "Builder differences are EXPECTED — GIIGNL lists the conversion "
                              "yard / an abbreviation; the backend keeps the original builder."))

    # --- Reclassify ------------------------------------------------------------
    rcols = ["giignl_name", "giignl_ex_names", "live_sheet_row", "backend_name",
             "backend_type", "cap_giignl", "cap_backend", "cap_agree",
             "giignl_owner", "giignl_terminal", "giignl_status", "suggested_action"]
    rrows = []
    for r in rec["reclassify"]:
        g, be, df = r["giignl"], r["backend"], r["diffs"]
        rrows.append({
            "giignl_name": g["vessel_name"],
            "giignl_ex_names": ", ".join(g.get("ex_names", [])),
            "live_sheet_row": be["sheet_row"], "backend_name": be["name"],
            "backend_type": be["vessel_type"],
            "cap_giignl": df["capacity"]["giignl_storage_m3"],
            "cap_backend": df["capacity"]["backend_capacity"],
            "cap_agree": _yn(df["capacity"]["agree"]),
            "giignl_owner": g.get("vessel_owner", ""),
            "giignl_terminal": g.get("location_raw", ""),
            "giignl_status": g.get("location_status", ""),
            "suggested_action": ("GIIGNL lists as FSRU; confirm and reclassify backend "
                                 "Vessel type (check 'candidate conversion' status before changing)"),
        })
    _table_sheet(wb, "Reclassify", rcols, rrows, fills=[FILL_YELLOW] * len(rrows),
                 widths={"A": 26, "B": 18, "D": 20, "I": 22, "J": 24, "L": 44})

    # --- Manual_pairing --------------------------------------------------------
    pcols = ["giignl_name", "live_sheet_row", "backend_name", "backend_type",
             "cap_giignl", "cap_backend", "owner_giignl", "owner_backend", "note"]
    prows = []
    for r in rec["manual"]:
        g, be, df = r["giignl"], r["suggested_backend"], r["diffs"]
        prows.append({
            "giignl_name": g["vessel_name"], "live_sheet_row": be["sheet_row"],
            "backend_name": be["name"], "backend_type": be["vessel_type"],
            "cap_giignl": df["capacity"]["giignl_storage_m3"],
            "cap_backend": df["capacity"]["backend_capacity"],
            "owner_giignl": df["owner"]["giignl"], "owner_backend": df["owner"]["backend"],
            "note": ("no name match; capacity+owner corroborate. Backend Name may be a "
                     "placeholder/defect (e.g. 'Hoegh' for 'Höegh Esperanza') — confirm "
                     "the pairing and fix the backend Name."),
        })
    _table_sheet(wb, "Manual_pairing", pcols, prows, fills=[FILL_YELLOW] * len(prows),
                 widths={"A": 24, "C": 20, "G": 20, "H": 20, "I": 60})

    # --- Backend_only ----------------------------------------------------------
    bcols = ["live_sheet_row", "backend_name", "imo", "capacity", "delivery_year",
             "classification", "note"]
    brows = [{
        "live_sheet_row": be["sheet_row"], "backend_name": be["name"], "imo": be["imo"],
        "capacity": be["capacity"], "delivery_year": be["delivery_year"],
        "classification": "expected (on-order / idle / spot)",
        "note": "GIIGNL is in-service only; verify it's not an FSU reclassification.",
    } for be in rec["backend_only"]]
    _table_sheet(wb, "Backend_only", bcols, brows, fills=[FILL_GRAY] * len(brows),
                 widths={"B": 26, "F": 30, "G": 50})

    # --- FSU_exclusions --------------------------------------------------------
    fcols = ["live_sheet_row", "name", "capacity", "owner", "note"]
    frows = [{
        "live_sheet_row": be["sheet_row"], "name": be["name"],
        "capacity": be["capacity"], "owner": be["owner"],
        "note": "FSU = storage-only, out of scope (inclusion_criteria.md).",
    } for be in rec["backend_fsu"]]
    _table_sheet(wb, "FSU_exclusions", fcols, frows, fills=[FILL_GRAY] * len(frows),
                 widths={"B": 26, "D": 24, "E": 48},
                 header_note=("Known exclusions so they don't resurface as gaps. NOTE: the "
                              "small-scale Candidates_to_add (e.g. Torman vs backend FSU "
                              "'Torman II') also pend FSRU-vs-small-scale/FSU adjudication."))

    # --- GIIGNL_orderbook (Phase B reference) ----------------------------------
    ocols = ["expected_year", "vessel_name", "storage_m3", "ccs", "owner", "builder"]
    orows = [{
        "expected_year": o.get("expected_year"), "vessel_name": o["vessel_name"],
        "storage_m3": o.get("storage_m3"), "ccs": o.get("ccs"),
        "owner": o.get("vessel_owner", ""), "builder": o.get("builder", ""),
    } for o in rec.get("orderbook", [])]
    _table_sheet(wb, "GIIGNL_orderbook", ocols, orows,
                 widths={"B": 24, "E": 22, "F": 24},
                 header_note="GIIGNL future deliveries — Phase B (on-order coverage), not added here.")

    # --- QA_review -------------------------------------------------------------
    qa_rows = [
        {"item": "Join method", "detail": "vessel name over {current} ∪ {ex_names}, "
         "normalize_vessel_name (diacritics + parentheticals folded). GIIGNL has no IMO."},
        {"item": "Corroborator", "detail": "capacity (storage m³ ↔ backend Capacity cbm); "
         f"agreed on all {s['matched']} matched pairs."},
        {"item": "Builder caveat", "detail": "informational only — GIIGNL lists the conversion "
         "yard for converted units; the backend keeps the original builder."},
        {"item": "Manual-match rule", "detail": "no-name suggestions require BOTH capacity "
         "agreement AND owner-tag overlap (guards coincidental capacity pairs)."},
        {"item": "Source rule", "detail": "GIIGNL is a comparison artifact, NOT a citable [ref]. "
         "No [ref] cells pre-filled; attach a primary source via url_verifier.py before promoting."},
        {"item": "Backend safety", "detail": "additive candidate workbook only; the backend is "
         "never auto-edited. Promote via the apply pipeline (digest → apply → verify)."},
        {"item": "URL verification gate", "detail": "§3.8 applies to any URL added during "
         "promotion research — not exercised here (no URLs cited)."},
    ]
    _table_sheet(wb, "QA_review", ["item", "detail"], qa_rows,
                 widths={"A": 24, "B": 110})

    out_path = _resolve_out_path(args.out, "lng_carrier_fsru_reconciliation.xlsx")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"  Wrote {out_path}")
    return out_path


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--mode", choices=["ref_fill", "discovery", "data_fill", "fix", "fsru"], required=True)
    p.add_argument("--rows", help="Row range for ref_fill, e.g. '1170-1190'")
    p.add_argument("--citations", help="Path to citations JSON (ref_fill mode)")
    p.add_argument("--candidates", help="Path to candidates JSON (discovery mode)")
    p.add_argument("--fills", help="Path to fills JSON (data_fill mode)")
    p.add_argument("--fix", help="Path to corrections JSON (fix mode)")
    p.add_argument("--reconcile", help="Path to FSRU reconciliation JSON (fsru mode; "
                                       "default work/fsru_reconcile.json)")
    p.add_argument("--base", help="Optional reviewed corrected-rows CSV to build on "
                                  "instead of the live (possibly corrupted) backend row (fix mode)")
    p.add_argument("--backend", default=str(backend_csv_path()),
                   help="Path to backend CSV (default: <work_dir>/backend.csv)")
    p.add_argument("--out",
                   help="Output path. Either a directory (the canonical xlsx "
                        "filename will be appended) or an explicit .xlsx file. "
                        "For batches, pass batches/<date>_<mode>_<scope>/.")
    args = p.parse_args()

    if args.mode == "ref_fill":
        if not args.citations:
            p.error("--citations required for ref_fill mode")
        build_ref_fill(args)
    elif args.mode == "discovery":
        if not args.candidates:
            p.error("--candidates required for discovery mode")
        build_discovery(args)
    elif args.mode == "data_fill":
        if not args.fills:
            p.error("--fills required for data_fill mode")
        build_data_fill(args)
    elif args.mode == "fsru":
        if not args.reconcile:
            args.reconcile = str(work_dir() / "fsru_reconcile.json")
        build_fsru(args)
    else:
        if not args.fix:
            p.error("--fix required for fix mode")
        build_fix(args)


if __name__ == "__main__":
    main()
